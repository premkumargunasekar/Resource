#!/usr/bin/python

from ansible.module_utils.basic import AnsibleModule
import ipaddress
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
import os
import tempfile
import hashlib
from datetime import datetime


# ==========================================================
# Utility Functions
# ==========================================================

def safe_string(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def sort_cidrs(cidr_list):
    return sorted(
        cidr_list,
        key=lambda x: int(ipaddress.ip_network(x, strict=False).network_address)
    )


def classify_vpc(network):
    name = str(network).lower()

    if "psa" in name:
        return "PSA"
    elif "inspection" in name:
        return "Inspection"
    elif "interconnect" in name:
        return "Interconnect"
    elif "host" in name or "shared" in name:
        return "Host Shared"
    else:
        return "Other"


def file_hash(path):
    if not os.path.exists(path):
        return None
    sha = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            sha.update(chunk)
    return sha.hexdigest()


# ==========================================================
# Core Logic
# ==========================================================

def generate_report(module):

    reserved_blocks = module.params["reserved_blocks"]
    subnets = module.params["subnets"]
    base_output = module.params["output_file"]
    threshold = module.params["utilization_threshold"]

    if threshold < 0 or threshold > 100:
        module.fail_json(msg="utilization_threshold must be between 0 and 100")

    today_str = datetime.now().strftime("%Y%m%d")
    base, ext = os.path.splitext(base_output)
    if not ext:
        ext = ".xlsx"
    output_file = f"{base}_{today_str}{ext}"

    required_keys = ["cidr", "region", "gcp_region", "env"]

    wb = Workbook()
    ws = wb.active
    ws.title = "IP Capacity Report"

    headers = [
        "Region", "GCP Region", "Env", "Reserved",
        "PSA VPC", "Inspection VPC",
        "Interconnect VPC", "Host Shared VPC",
        "Other VPC",
        "Available Ranges", "Used IPs", "Free IPs", "Util %"
    ]

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    row_num = 2

    # Sort reserved blocks for deterministic output
    reserved_blocks = sorted(
        reserved_blocks,
        key=lambda x: ipaddress.ip_network(x["cidr"], strict=False).network_address
    )

    for block in reserved_blocks:

        for key in required_keys:
            if key not in block:
                module.fail_json(msg=f"Missing required key '{key}'")

        reserved_net = ipaddress.ip_network(block["cidr"], strict=False)

        vpc_groups = defaultdict(list)

        for subnet in subnets:
            subnet_cidr = subnet.get("ipCidrRange")
            network_name = subnet.get("network", "")

            if not subnet_cidr:
                continue

            try:
                subnet_net = ipaddress.ip_network(subnet_cidr, strict=False)
            except Exception:
                module.warn(f"Skipping invalid subnet CIDR: {subnet_cidr}")
                continue

            if subnet_net.subnet_of(reserved_net):
                vpc_groups[classify_vpc(network_name)].append(subnet_cidr)

        # Deterministic ordering
        for k in vpc_groups:
            vpc_groups[k] = sort_cidrs(vpc_groups[k])

        used_networks = sorted(
            [ipaddress.ip_network(c, strict=False)
             for v in vpc_groups.values() for c in v],
            key=lambda x: int(x.network_address)
        )

        used_ips = sum(n.num_addresses for n in used_networks)
        total_ips = reserved_net.num_addresses
        free_ips = total_ips - used_ips
        util = round((used_ips / total_ips) * 100, 2) if total_ips else 0

        free_ranges = [reserved_net]

        for used in used_networks:
            new_free = []
            for free in free_ranges:
                if used.overlaps(free):
                    new_free.extend(free.address_exclude(used))
                else:
                    new_free.append(free)
            free_ranges = new_free

        free_ranges = sorted(
            free_ranges,
            key=lambda x: x.num_addresses,
            reverse=True
        )

        available = ", ".join(str(r) for r in free_ranges[:3])

        ws.append([
            safe_string(block["region"]),
            safe_string(block["gcp_region"]),
            safe_string(block["env"]),
            safe_string(block["cidr"]),
            ", ".join(vpc_groups.get("PSA", [])),
            ", ".join(vpc_groups.get("Inspection", [])),
            ", ".join(vpc_groups.get("Interconnect", [])),
            ", ".join(vpc_groups.get("Host Shared", [])),
            ", ".join(vpc_groups.get("Other", [])),
            available,
            used_ips,
            free_ips,
            util
        ])

        row_num += 1

    red_fill = PatternFill(start_color="FFC7CE",
                           end_color="FFC7CE",
                           fill_type="solid")

    green_fill = PatternFill(start_color="C6EFCE",
                             end_color="C6EFCE",
                             fill_type="solid")

    red_rule = FormulaRule(formula=[f"$M2>{threshold}"], fill=red_fill)
    green_rule = FormulaRule(formula=[f"$M2<={threshold}"], fill=green_fill)

    ws.conditional_formatting.add(f"A2:M{row_num}", red_rule)
    ws.conditional_formatting.add(f"A2:M{row_num}", green_rule)

    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            temp_path = tmp.name
        wb.save(temp_path)
        wb.close()

        old_hash = file_hash(output_file)
        new_hash = file_hash(temp_path)

        if old_hash == new_hash:
            os.remove(temp_path)
            module.exit_json(changed=False, output_file=output_file)

        os.replace(temp_path, output_file)

    except Exception as e:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)
        module.fail_json(msg=str(e))

    module.exit_json(changed=True, output_file=output_file)


def main():

    module = AnsibleModule(
        argument_spec=dict(
            reserved_blocks=dict(type='list', elements='dict', required=True),
            subnets=dict(type='list', elements='dict', required=True),
            output_file=dict(type='str', required=True),
            utilization_threshold=dict(type='int', default=60)
        ),
        supports_check_mode=True
    )

    if module.check_mode:
        module.exit_json(changed=False)

    generate_report(module)


if __name__ == "__main__":
    main()
